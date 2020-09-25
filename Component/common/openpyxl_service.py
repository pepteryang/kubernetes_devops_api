from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl import Workbook
from io import BytesIO
import datetime
from django.db.models import Model


def get_format_value(self, obj, field):
    """将对象属性格式化"""
    attr = getattr(obj, field)
    value = attr
    # 外键格式化
    if isinstance(attr, Model):
        value = attr.__str__()
    # 时间类型格式化
    if isinstance(attr, datetime.datetime):
        value = attr.strftime('%Y-%m-%d %H:%M:%S')
    # 其他的暂时不做转换
    return value

def handle_data_to_excl(export_data, report_name):
    file_io = BytesIO()
    wb = Workbook()
    ws = wb.worksheets[0]
    # 水平居中, 垂直居中
    alignment_style = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    # 设置ws的名称
    ws.title = report_name
    # 合并单元格 合并第一行的前2列
    x = 1

    # 定义Border边框样式
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    border_style = Border(left=left, right=right, top=top, bottom=bottom)

    # 单元格合并
    # ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(export_data['head_list']))
    # ws.cell(row=1, column=1).value = report_name

    # # 第2行写入值
    # now_time = datetime.datetime.now().strftime('%Y%m%d %H%M')
    # # 格式化时间为中文年月日
    # now_time = now_time[:4] + '年' + now_time[4:6] + '月' + now_time[6:8] +
    # '日' + now_time[8:11] + '时' + now_time[11:13] + '分'
    #
    # ws.cell(row=2, column=1).value = '报表导出时间：{}'.format(now_time)
    # # ws.cell(row=2, column=3).value = '后台系统版本：Light_OS_Back_ver1.04'
    # # ws.cell(row=2, column=5).value = '终端系统版本:Light_OS_Front_ver1.06'
    # ws.cell(row=2, column=3).value = '报表导出位置：{}'.format(report_name)
    # 关闭默认灰色网格线
    ws.sheet_view.showGridLines = False
    # 第一行行高设置为22
    ws.row_dimensions[1].height = 22

    border = Border(left=Side(style='thin', color='FF000000'),
                    right=Side(style='thin', color='FF000000'),
                    top=Side(style='thin', color='FF000000'),
                    bottom=Side(style='thin', color='FF000000'),
                    diagonal=Side(style='thin', color='FF000000'),
                    diagonal_direction=0, outline=Side(style='medium', color='FF000000'),
                    vertical=Side(style='thin', color='FF000000'),
                    horizontal=Side(style='thin', color='FF000000'))

    # 对单元格进行填充
    fill_heading = PatternFill('solid', fgColor='BFBFBF')  # 灰色
    fill = PatternFill('solid', fgColor='FF9999')  # 亮粉

    # 表头
    for head in range(1, len(export_data['head_list']) + 1):
        # 第三行写入值
        ws.cell(row=1, column=head).value = export_data['head_list'][head - 1]
        # 设置文本水平居中, 垂直居中
        ws.cell(row=1, column=head).alignment = alignment_style
        #  设置字体加粗
        ws.cell(row=1, column=head).font = Font(bold=True, size=9)
        # 背景颜色
        ws.cell(row=1, column=head).fill = PatternFill(fill_type='solid', fgColor='EE9A49')
        # 边框样式
        ws.cell(row=1, column=head).border = border_style

    # 表数据
    excel_row = 2
    for data in export_data['data_list']:
        excel_col = 1
        for i in range(0, len(data)):
            if i % 2:
                # 数据
                ws.cell(row=excel_row, column=excel_col).value = list(data)[i]
                # 颜色
                ws.cell(row=excel_row, column=excel_col).fill = fill_heading
                # 边框
                ws.cell(row=excel_row, column=excel_col).border = border
            else:
                ws.cell(row=excel_row, column=excel_col).value = list(data)[i]
                ws.cell(row=excel_row, column=excel_col).fill = fill
                ws.cell(row=excel_row, column=excel_col).border = border
            s = chr(excel_col + 64) + str(excel_row)
            # wrap_text=True 打开自动换行
            ws[s].alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions.values()
            ws[s].font = Font(size=10)
            excel_col = excel_col + 1
        excel_row = excel_row + 1
    wb.save(file_io)
    return file_io

